import os
import openpyxl
import math



print('Data will be read from excel in the "Dataset" folder.')
print('First column will be treated as X and the second will be assumed Y. Also first row will be title.\n')
print('The bias is Y = a2 X ^ b2.\nYou will give MAX value of b2 (like 50).')
print('The app will first start to calculate each b2 and reports each value results.')
print('Then it will compare and report best found b2 from 1 to your input (each step adds 0.1 to the former value).')
print('At the end the app will calculate b2 by itself and reports the best found b2 value.')


dataframe = openpyxl.load_workbook(os.getcwd() + r'\Dataset\PreparedDataset.xlsx') 
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
#for row in range(0, dataframe1.max_row):
#    for col in dataframe1.iter_cols(1, dataframe1.max_column):
#        print(col[row].value)

maxB = float(input('Please enter maximum value of b2: '))
print('\n\n\n')


XStarMean = 0
YStarMean = 0

for row in range(1, dataframe1.max_row): 
    for col in dataframe1.iter_cols(1, 1):
        XStarMean+=math.log10(col[row].value)

    for col in dataframe1.iter_cols(dataframe1.max_column, dataframe1.max_column):
        YStarMean+=math.log10(col[row].value)

XStarMean/=(dataframe1.max_row-1)
YStarMean/=(dataframe1.max_row-1) 

rValuesList = []

for b2 in range(1,int((maxB)*10)+1):
    b2 /=10
    print(f'for b2 = {b2}:')
    print(f'Y = a2 X ^ {b2}')
    print(f'log(Y) = log(a2 X ^ {b2})')
    print(f'log(Y) = log(a2) + {b2}log(X)')
    print(f'Y* = log(Y)')
    print(f'a0 = log(a2)')
    print(f'a1 = {b2}')
    print(f'X* = log(X)')
    print(f'Y* = a0 + {b2}X*')
    print(f'X*\u0304 = {XStarMean}')
    print(f'Y*\u0304 = {YStarMean}')    
    print(f'a0 = Y*\u0304 - {b2}X*\u0304')
    a0 = YStarMean - (b2*XStarMean)
    print(f'a0 = {YStarMean} - ( {b2} * {XStarMean}) = {a0}')
    print(f'log(Y) = {a0} + {b2}log(X)')
    a2 = math.pow(10,a0)
    print(f'a2 = math.pow(10,a0) = math.pow(10,{a0}) = {a2}\n')
    print(f'Y = {a2} X ^ {b2}')
    st = 0
    sr = 0

     
    for row in range(1, dataframe1.max_row): 
        foundY = 0
        for col in dataframe1.iter_cols(1, 1):
            foundY= a2 * math.pow(col[row].value,b2)

        for col in dataframe1.iter_cols(dataframe1.max_column, dataframe1.max_column):
            st+=math.pow(col[row].value-YStarMean,2)
            sr+=math.pow(col[row].value-foundY,2) 

    print(f'st = {st}')
    print(f'sr = {sr}')
    r2 = (st-sr)/st
    print(f'r^2 = (st-sr)/st = {r2}')
    if r2<0:
        print('r2 is negative - EFTEZAHE!')
        r = r2
    else:
        r = math.sqrt(r2)

    print(f'r = {r}')
    rValuesList.append(1-r)
    print('\n\n\n')

    
bestR =min(rValuesList) 
bestB2Value = (rValuesList.index(bestR) + 1)/10

print(f'best b2 value was {bestB2Value} with r = {1-bestR}')


print('\n\n\n')
print('Now we will calculate b2:')
print(f'Y = a2 X ^ b2')
print(f'log(Y) = log(a2 X ^ b2)')
print(f'log(Y) = log(a2) + b2 * log(X)')
print(f'Y* = log(Y)')
print(f'a0 = log(a2)')
print(f'a1 = b2')
print(f'X* = log(X)')
print(f'Y* = a0 + a1 * X*')
print(f'X*\u0304 = {XStarMean}')
print(f'Y*\u0304 = {YStarMean}')    




sumOfXStarInYStar = 0
sumOfXStar = 0
sumOfXStarPow2 = 0
sumOfYStar = 0
     
for row in range(1, dataframe1.max_row): 
    foundY = 0
    for col in dataframe1.iter_cols(1, 1):
        xStar= math.log10(col[row].value)
        sumOfXStar+=xStar
        sumOfXStarPow2+=math.pow(xStar,2)

    for col in dataframe1.iter_cols(dataframe1.max_column, dataframe1.max_column):
        yStar= math.log10(col[row].value)
        sumOfYStar+=yStar

    sumOfXStarInYStar+= xStar * yStar
    
countOfData = (dataframe1.max_row-1)
a1 = ((countOfData * sumOfXStarInYStar) - (sumOfXStar * sumOfYStar))/((countOfData * sumOfXStarPow2) - math.pow(sumOfXStar,2))
print(f'a1 = (({countOfData} * {sumOfXStarInYStar}) - ({sumOfXStar} * {sumOfYStar}))/(({countOfData} * {sumOfXStarPow2}) - {math.pow(sumOfXStar,2)}) = {a1}')
 
print(f'a0 = Y*\u0304 - a1 X*\u0304')
a0 = YStarMean - (a1*XStarMean)
print(f'a0 = {YStarMean} - ( {a1} * {XStarMean}) = {a0}')
print(f'log(Y) = {a0} + {a1}log(X)')
a2 = math.pow(10,a0)
print(f'a2 = math.pow(10,a0) = math.pow(10,{a0}) = {a2}\n')
print(f'Y = {a2} X ^ {a1}')
st = 0
sr = 0

     
for row in range(1, dataframe1.max_row): 
    foundY = 0
    for col in dataframe1.iter_cols(1, 1):
        foundY= a2 * math.pow(col[row].value,a1)

    for col in dataframe1.iter_cols(dataframe1.max_column, dataframe1.max_column):
        st+=math.pow(col[row].value-YStarMean,2)
        sr+=math.pow(col[row].value-foundY,2) 

print(f'st = {st}')
print(f'sr = {sr}')
r2 = (st-sr)/st
print(f'r^2 = (st-sr)/st = {r2}')
if r2<0:
    print('r2 is negative - EFTEZAHE!')
    r = r2
else:
    r = math.sqrt(r2)

print(f'r = {r}')

print('\n\n\n\n\n\n') 
print(f'So the best b2 was actually {a1}!')


